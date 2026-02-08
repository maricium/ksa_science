"""
Excel Reader
Consolidated utilities for finding and reading Excel files from Lesson Resources

This module provides reusable functions for:
- Finding Excel files for specific units/lessons
- Reading Excel files with pandas or openpyxl
- Locating columns (keywords, vocabulary, lesson codes)
- Extracting data from specific rows/columns
- Parsing keyword definitions and lists

Usage:
    from readers.excel_reader import find_excel_file, get_keywords_from_excel

    # Find an Excel file for a unit
    excel_file = find_excel_file("B3.2", "B3.2.7")

    # Get keywords for a lesson
    keywords = get_keywords_from_excel("B3.2.7", "B3.2")
"""

from __future__ import annotations

import logging
import re
from functools import lru_cache
from pathlib import Path

import openpyxl
import pandas as pd

logger = logging.getLogger("readers.excel_reader")


@lru_cache(maxsize=64)
def find_excel_file(unit: str, lesson_code: str | None = None) -> Path | None:
    """
    Find Excel file for a given unit code.

    Args:
        unit: Unit code (e.g., "B3.2", "C4.1")
        lesson_code: Optional lesson code for additional context

    Returns:
        Path to Excel file or None if not found
    """
    # Try to find Excel files in Lesson Resources/Unit Guidance
    lesson_resources = Path("Lesson Resources")
    if not lesson_resources.exists():
        lesson_resources = Path("..") / "Lesson Resources"
    if not lesson_resources.exists():
        lesson_resources = Path("../..") / "Lesson Resources"

    # Search for Excel files in Unit Guidance folders
    excel_files = list(lesson_resources.rglob(f"**/Unit Guidance/**/{unit}*.xlsx"))
    if not excel_files:
        excel_files = list(lesson_resources.rglob(f"**/Unit Guidance/**/*{unit}*.xlsx"))

    # Also try the old Excel_Files location as fallback
    if not excel_files:
        excel_file = Path(f"Excel_Files/{unit}_Unit_Overview.xlsx")
        if excel_file.exists():
            excel_files = [excel_file]
        else:
            excel_file = Path(f"../Excel_Files/{unit}_Unit_Overview.xlsx")
            if excel_file.exists():
                excel_files = [excel_file]

    # Also check for C3.1 format in old location
    if not excel_files and unit.startswith("C3"):
        excel_file = Path("Excel_Files/C3.1 Unit Overview and Planning Proforma.xlsx")
        if excel_file.exists():
            excel_files = [excel_file]
        else:
            excel_file = Path("../Excel_Files/C3.1 Unit Overview and Planning Proforma.xlsx")
            if excel_file.exists():
                excel_files = [excel_file]

    if excel_files:
        return excel_files[0]  # Use first match

    return None


def find_keyword_column(worksheet, header_rows: list[int] = None) -> tuple[int | None, int | None, int | None]:
    """
    Find the keyword/vocabulary column index in an Excel worksheet.

    Args:
        worksheet: openpyxl worksheet object
        header_rows: List of row numbers to check for headers (1-indexed)

    Returns:
        Tuple of (keyword_col_idx, lesson_code_col_idx, header_row)
        All indices are 0-indexed for use with openpyxl
    """
    if header_rows is None:
        header_rows = [6, 5]
    keyword_col_idx = None
    lesson_code_col_idx = 0

    for header_row in header_rows:
        # openpyxl uses 1-indexed rows, so header_row is already correct
        for col_idx, cell in enumerate(worksheet[header_row], start=1):
            cell_val = str(cell.value).lower() if cell.value else ""
            if (
                "vocabulary" in cell_val
                or "literacy" in cell_val
                or ("keyword" in cell_val and "difficult" in cell_val)
            ):
                keyword_col_idx = col_idx - 1  # Convert to 0-indexed
                break

        if keyword_col_idx is not None:
            # Adjust lesson_code_col_idx based on header row
            if header_row == 5:
                lesson_code_col_idx = 1
            return keyword_col_idx, lesson_code_col_idx, header_row

    return None, None, None


def find_lesson_row(worksheet, lesson_code: str, lesson_code_col_idx: int, start_row: int = 6) -> int | None:
    """
    Find the row number for a specific lesson code in an Excel worksheet.

    Args:
        worksheet: openpyxl worksheet object
        lesson_code: Lesson code to find (e.g., "B3.2.7")
        lesson_code_col_idx: Column index (0-indexed) where lesson codes are stored
        start_row: Row number to start searching from (1-indexed)

    Returns:
        Row number (1-indexed) if found, None otherwise
    """
    for row in worksheet.iter_rows(min_row=start_row, max_row=worksheet.max_row):
        if len(row) > lesson_code_col_idx and row[lesson_code_col_idx].value:
            row_lesson_code = str(row[lesson_code_col_idx].value).strip()
            if row_lesson_code.upper() == lesson_code.upper():
                return row[lesson_code_col_idx].row
    return None


def get_lesson_row_data(worksheet, lesson_code: str, lesson_code_col_idx: int, start_row: int = 6) -> tuple | None:
    """
    Get the row data for a specific lesson code.

    Args:
        worksheet: openpyxl worksheet object
        lesson_code: Lesson code to find
        lesson_code_col_idx: Column index (0-indexed) where lesson codes are stored
        start_row: Row number to start searching from (1-indexed)

    Returns:
        Tuple of row values if found, None otherwise
    """
    for row in worksheet.iter_rows(values_only=True, min_row=start_row):
        row_lesson_code = (
            str(row[lesson_code_col_idx]).strip()
            if len(row) > lesson_code_col_idx and row[lesson_code_col_idx]
            else None
        )
        if row_lesson_code and row_lesson_code.upper() == lesson_code.upper():
            return row
    return None


def extract_keyword_definition(keyword_text: str, keyword: str) -> str | None:
    """
    Extract definition for a keyword from keyword text.

    Args:
        keyword_text: Text containing keywords and definitions
        keyword: The keyword to find definition for

    Returns:
        Definition string if found, None otherwise
    """
    keyword_lower = keyword.lower()
    for line in keyword_text.split("\n"):
        line_lower = line.lower()
        # Check if this line contains the keyword
        if keyword_lower in line_lower:
            # Try to extract definition (look for patterns like "X is...", "X: ...", "X means...")
            definition_match = re.search(rf"{re.escape(keyword)}\s*[:\-–]\s*(.+?)(?:\n|$)", line, re.IGNORECASE)
            if definition_match:
                return definition_match.group(1).strip()
            # Or if the line itself looks like a definition
            if "is" in line_lower or "means" in line_lower or "refers to" in line_lower:
                # Extract everything after the keyword
                parts = re.split(rf"{re.escape(keyword)}\s*[:\-–]?\s*", line, flags=re.IGNORECASE, maxsplit=1)
                if len(parts) > 1:
                    return parts[1].strip()
    return None


def get_keywords_from_cell(keyword_text: str) -> list[str]:
    """
    Extract list of keywords from a cell's text content.

    Args:
        keyword_text: Text containing keywords (comma or semicolon separated)

    Returns:
        List of keyword strings
    """
    if not keyword_text or keyword_text.lower() in ["none", "nan", ""]:
        return []

    keywords = []
    for line in keyword_text.split("\n"):
        keywords.extend([kw.strip() for kw in re.split(r"[,;]", line.strip()) if kw.strip()])
    return keywords


@lru_cache(maxsize=32)
def _read_excel_with_openpyxl_cached(excel_file_str: str) -> openpyxl.Workbook | None:
    """
    Cached version of read_excel_with_openpyxl (internal use).

    Args:
        excel_file_str: Path to Excel file as string (for hashing)

    Returns:
        openpyxl Workbook object or None if error
    """
    try:
        return openpyxl.load_workbook(excel_file_str, data_only=True)
    except Exception as e:
        logger.warning(f"Error reading Excel file with openpyxl: {e}")
        return None


def read_excel_with_openpyxl(excel_file: Path) -> openpyxl.Workbook | None:
    """
    Read Excel file using openpyxl (with caching).

    Args:
        excel_file: Path to Excel file

    Returns:
        openpyxl Workbook object or None if error
    """
    return _read_excel_with_openpyxl_cached(str(excel_file.absolute()))


@lru_cache(maxsize=64)
def _read_excel_with_pandas_cached(
    excel_file_str: str, sheet_name: str | None, header: int | None, skiprows: int | None
) -> pd.DataFrame | None:
    """
    Cached version of read_excel_with_pandas (internal use).
    Returns a copy to avoid mutable cache issues.

    Args:
        excel_file_str: Path to Excel file as string (for hashing)
        sheet_name: Optional sheet name
        header: Optional header row number
        skiprows: Optional number of rows to skip

    Returns:
        pandas DataFrame (copy) or None if error
    """
    try:
        kwargs = {}
        if sheet_name:
            kwargs["sheet_name"] = sheet_name
        if header is not None:
            kwargs["header"] = header
        if skiprows is not None:
            kwargs["skiprows"] = skiprows

        df = pd.read_excel(excel_file_str, **kwargs)
        # Return a copy to avoid mutable cache issues
        return df.copy() if df is not None else None
    except Exception as e:
        logger.warning(f"Error reading Excel file with pandas: {e}")
        return None


def read_excel_with_pandas(
    excel_file: Path, sheet_name: str | None = None, header: int | None = None, skiprows: int | None = None
) -> pd.DataFrame | None:
    """
    Read Excel file using pandas (with caching).
    Returns a copy of the cached DataFrame to avoid mutable cache issues.

    Args:
        excel_file: Path to Excel file
        sheet_name: Optional sheet name
        header: Optional header row number
        skiprows: Optional number of rows to skip

    Returns:
        pandas DataFrame (copy) or None if error
    """
    return _read_excel_with_pandas_cached(str(excel_file.absolute()), sheet_name, header, skiprows)


def find_sheet_name_for_unit(excel_file: Path, unit: str) -> str | None:
    """
    Find the appropriate sheet name for a given unit in an Excel file.

    Args:
        excel_file: Path to Excel file
        unit: Unit code (e.g., "C4.2")

    Returns:
        Sheet name if found, None otherwise
    """
    try:
        xl_file = pd.ExcelFile(excel_file)
        # Look for sheet with unit code in the name
        for sheet in xl_file.sheet_names:
            if unit in sheet:
                return sheet
        # If not found, return first sheet
        if xl_file.sheet_names:
            return xl_file.sheet_names[0]
    except Exception:
        pass
    return None


def get_lesson_data(lesson_code: str, unit: str | None = None) -> dict | None:
    """
    Extract lesson data (KNOW and DO) for a specific lesson code.

    This function handles multiple Excel file formats:
    - C3.1 format: Uses specific sheet name and header=3
    - C4.2 format: Uses sheet name detection and skiprows=3
    - Generic format: Tries different skip rows to find lesson code column

    Args:
        lesson_code: Lesson code (e.g., "C4.1.4" or "C3.1.3")
        unit: Unit code (e.g., "C4.1"). If None, extracted from lesson_code

    Returns:
        dict with 'code', 'title', 'know', 'do' or None if not found
    """
    # Determine which Excel file to use based on unit
    if not unit:
        unit = ".".join(lesson_code.split(".")[:2])  # Extract C4.1 from C4.1.4

    # Find Excel file using helper
    excel_file = find_excel_file(unit, lesson_code)
    if not excel_file or not excel_file.exists():
        logger.error(f"Excel file not found for unit {unit}")
        logger.error(f"Searched in: Lesson Resources/**/Unit Guidance/**/{unit}*.xlsx")
        return None

    logger.info(f"Reading Excel file: {excel_file.name}")
    logger.info(f"Unit: {unit}, Lesson code: {lesson_code}")

    # Try to determine sheet name for C4.2 files
    sheet_name = None
    if unit.startswith("C4.2"):
        sheet_name = find_sheet_name_for_unit(excel_file, "C4.2")

    # C3.1 files have different structure - use header=3, lesson codes in column 1
    if unit.startswith("C3"):
        try:
            df = read_excel_with_pandas(excel_file, sheet_name="C3.1 The Periodic Table", header=3)
            if df is not None:
                df = df.dropna(how="all")

                # Column 1 has lesson codes, Column 2 has title, Column 3 has "know" content
                for _idx, row in df.iterrows():
                    if len(row) > 1 and pd.notna(row.iloc[1]):
                        lesson_code_in_file = str(row.iloc[1]).strip()
                        if lesson_code_in_file == lesson_code:
                            # Column 2 has title, Column 3 has "know" content (bullet points)
                            title = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else lesson_code
                            know_content = str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else ""
                            return {
                                "code": lesson_code,
                                "title": title,
                                "know": know_content,
                                "do": str(row.iloc[4]).strip() if len(row) > 4 and pd.notna(row.iloc[4]) else "",
                            }
        except Exception as e:
            logger.warning(f"Error reading C3.1 file: {e}")
            import traceback

            traceback.print_exc()

    # Also check if lesson_code itself starts with C3 (in case unit wasn't set correctly)
    if lesson_code.startswith("C3") and not unit.startswith("C3"):
        try:
            df = read_excel_with_pandas(excel_file, sheet_name="C3.1 The Periodic Table", header=3)
            if df is not None:
                df = df.dropna(how="all")

                for _idx, row in df.iterrows():
                    if len(row) > 1 and pd.notna(row.iloc[1]):
                        lesson_code_in_file = str(row.iloc[1]).strip()
                        if lesson_code_in_file == lesson_code:
                            title = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else lesson_code
                            know_content = str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else ""
                            return {
                                "code": lesson_code,
                                "title": title,
                                "know": know_content,
                                "do": str(row.iloc[4]).strip() if len(row) > 4 and pd.notna(row.iloc[4]) else "",
                            }
        except Exception:
            pass

    # Try reading with sheet name if we found one (for C4.2 files)
    if sheet_name:
        try:
            df = read_excel_with_pandas(excel_file, sheet_name=sheet_name, skiprows=3)
            if df is not None:
                for _idx, row in df.iterrows():
                    # C4.2 format: lesson code is in column 1
                    if len(row) > 1 and pd.notna(row.iloc[1]):
                        code_col1 = str(row.iloc[1]).strip()
                        if code_col1 == lesson_code:
                            return {
                                "code": lesson_code,
                                "title": str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else "",
                                "know": str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else "",
                                "do": str(row.iloc[5]).strip() if len(row) > 5 and pd.notna(row.iloc[5]) else "",
                            }
        except Exception:
            pass

    # Try different skip rows to find the data (for C4 files and fallback)
    for skip in range(10):
        try:
            df = read_excel_with_pandas(excel_file, skiprows=skip, sheet_name=sheet_name)
            if df is None:
                continue

            first_col = str(df.columns[0]).lower().strip()

            if "lesson" in first_col and "code" in first_col:
                # Check both column 0 and column 1 for lesson codes
                for _idx, row in df.iterrows():
                    # Try column 1 first (C4.2 format - lesson code is in column 1)
                    if len(row) > 1 and pd.notna(row.iloc[1]):
                        code_col1 = str(row.iloc[1]).strip()
                        if code_col1 == lesson_code:
                            return {
                                "code": lesson_code,
                                "title": str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else "",
                                "know": str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else "",
                                "do": str(row.iloc[5]).strip() if len(row) > 5 and pd.notna(row.iloc[5]) else "",
                            }
                    # Try column 0 (other formats)
                    if len(row) > 0 and pd.notna(row.iloc[0]):
                        code_col0 = str(row.iloc[0]).strip()
                        if code_col0 == lesson_code:
                            return {
                                "code": lesson_code,
                                "title": str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else "",
                                "know": str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else "",
                                "do": str(row.iloc[4]).strip() if len(row) > 4 and pd.notna(row.iloc[4]) else "",
                            }
        except Exception:
            continue

    return None
